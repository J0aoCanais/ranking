import os
from datetime import datetime
from typing import Tuple
from django.conf import settings
from openpyxl import Workbook
from openpyxl import load_workbook
from .models import Person


def export_persons_to_local_excel_and_delete() -> Tuple[int, str]:
	"""Export all persons to a local XLSX file and delete them.
	Returns: (count, relative_url)
	The file is saved in MEDIA_ROOT/exports/persons-YYYYmmdd-HHMMSS.xlsx
	relative_url is under MEDIA_URL for download.
	"""
	persons = list(Person.objects.all().order_by('id'))
	if not persons:
		# still create an empty file for traceability
		pass

	wb = Workbook()
	ws = wb.active
	ws.title = 'Pessoas'
	# header
	ws.append(['Primeiro Nome', 'Último Nome', 'Álcool', 'Data Adicionado'])
	for p in persons:
		ws.append([p.primeiro_nome, p.ultimo_nome, p.alcool, p.data_adicionado.strftime('%Y-%m-%d %H:%M:%S')])

	# ensure directory
	exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
	os.makedirs(exports_dir, exist_ok=True)
	# Fixed filename to overwrite on each export
	filename = 'pessoas.xlsx'
	file_path = os.path.join(exports_dir, filename)
	wb.save(file_path)

	# delete persons
	if persons:
		Person.objects.filter(id__in=[p.id for p in persons]).delete()

	relative_url = f"{settings.MEDIA_URL.rstrip('/')}/exports/{filename}"
	return (len(persons), relative_url)


def _excel_file_paths() -> Tuple[str, str]:
	exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
	filename = 'pessoas.xlsx'
	file_path = os.path.join(exports_dir, filename)
	url = f"{settings.MEDIA_URL.rstrip('/')}/exports/{filename}"
	return file_path, url


def clear_local_excel_file() -> str:
	"""Create or reset pessoas.xlsx with only the header row. Returns the URL."""
	file_path, url = _excel_file_paths()
	os.makedirs(os.path.dirname(file_path), exist_ok=True)
	wb = Workbook()
	ws = wb.active
	ws.title = 'Pessoas'
	ws.append(['Primeiro Nome', 'Último Nome', 'Álcool', 'Data Adicionado'])
	wb.save(file_path)
	return url


def get_local_excel_info() -> dict:
	"""Return info about pessoas.xlsx: exists, rows (excluding header), url."""
	file_path, url = _excel_file_paths()
	exists = os.path.exists(file_path)
	rows = 0
	if exists:
		wb = load_workbook(file_path)
		ws = wb.active
		# count rows excluding header if there is at least one row
		total = ws.max_row or 0
		rows = max(total - 1, 0)
	return {'exists': exists, 'rows': rows, 'url': url}
